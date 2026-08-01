from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LogCallback = Callable[[str], None]


@dataclass
class InventoryItem:
    kind: str
    name: str
    stem: str
    extension: str
    parent_name: str
    relative_path: str
    full_path: str
    size_bytes: int | str
    size_display: str
    modified_at: str
    created_at: str
    direct_file_count: int | str = ""
    direct_folder_count: int | str = ""


@dataclass
class InventoryResult:
    rows: list[InventoryItem] = field(default_factory=list)
    ext_summary: dict[str, dict] = field(default_factory=dict)
    folder_summary: dict[str, dict] = field(default_factory=dict)
    errors: list[dict] = field(default_factory=list)


def normalize_ext(ext: str) -> str:
    ext = (ext or "").strip().lower()
    if not ext:
        return ""
    if not ext.startswith("."):
        ext = "." + ext
    return ext


def parse_extension_filter(text: str) -> set[str]:
    text = (text or "").strip()
    if not text:
        return set()

    parts: list[str] = []
    for part in text.replace(";", ",").replace(" ", ",").split(","):
        part = part.strip()
        if part:
            parts.append(part)
    return {normalize_ext(part) for part in parts if normalize_ext(part)}


def format_datetime(timestamp) -> str:
    try:
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def format_size(num_bytes) -> str:
    try:
        value = int(num_bytes)
    except Exception:
        return ""

    if value >= 1024 * 1024 * 1024:
        return f"{value / (1024 * 1024 * 1024):,.2f} GB"
    if value >= 1024 * 1024:
        return f"{value / (1024 * 1024):,.2f} MB"
    if value >= 1024:
        return f"{value / 1024:,.2f} KB"
    return f"{value:,} B"


def get_downloads_folder() -> Path:
    downloads = Path.home() / "Downloads"
    if downloads.is_dir():
        return downloads
    return Path.home()


def default_excel_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return get_downloads_folder() / f"파일_현황표_{timestamp}.xlsx"


def safe_excel_sheet_title(name: str) -> str:
    invalid = '[]:*?/\\'
    result = name or "Sheet"
    for char in invalid:
        result = result.replace(char, "_")
    return result[:31] or "Sheet"


def is_office_temp_file(name: str) -> bool:
    return name.startswith("~$")


def count_direct_children(folder_path: str | Path) -> tuple[int, int]:
    file_count = 0
    folder_count = 0
    try:
        with os.scandir(folder_path) as iterator:
            for entry in iterator:
                try:
                    if entry.is_dir(follow_symlinks=False):
                        folder_count += 1
                    elif entry.is_file(follow_symlinks=False):
                        file_count += 1
                except Exception:
                    pass
    except Exception:
        pass
    return file_count, folder_count


def collect_inventory(
    root_folder: str | Path,
    include_subfolders: bool,
    target_mode: str,
    ext_filter: set[str],
    exclude_office_temp: bool,
    log_func: LogCallback | None = None,
) -> InventoryResult:
    root = Path(root_folder).resolve()
    result = InventoryResult()

    include_files = target_mode in ("파일만", "파일+폴더")
    include_folders = target_mode in ("폴더만", "파일+폴더")

    def log(message: str) -> None:
        if log_func is not None:
            log_func(message)

    def should_skip_name(name: str) -> bool:
        return exclude_office_temp and is_office_temp_file(name)

    def add_folder_summary(parent_path: Path, kind: str, size_bytes: int = 0) -> None:
        try:
            rel_parent = os.path.relpath(parent_path, root)
        except Exception:
            rel_parent = str(parent_path)
        if rel_parent == ".":
            rel_parent = "(기준 폴더)"

        item = result.folder_summary.setdefault(
            rel_parent,
            {
                "relative_path": rel_parent,
                "full_path": str(parent_path),
                "file_count": 0,
                "folder_count": 0,
                "total_size_bytes": 0,
            },
        )

        if kind == "파일":
            item["file_count"] += 1
            item["total_size_bytes"] += size_bytes
        elif kind == "폴더":
            item["folder_count"] += 1

    def add_ext_summary(extension: str, size_bytes: int) -> None:
        ext_key = extension.lower() if extension else "(확장자 없음)"
        item = result.ext_summary.setdefault(
            ext_key,
            {
                "extension": ext_key,
                "file_count": 0,
                "total_size_bytes": 0,
            },
        )
        item["file_count"] += 1
        item["total_size_bytes"] += size_bytes

    def add_file(path: Path) -> None:
        try:
            name = path.name
            if should_skip_name(name):
                return

            extension = normalize_ext(path.suffix)
            if ext_filter and extension not in ext_filter:
                return

            stat = path.stat()
            parent = path.parent
            size_bytes = stat.st_size
            relative_path = os.path.relpath(path, root)

            result.rows.append(
                InventoryItem(
                    kind="파일",
                    name=name,
                    stem=path.stem,
                    extension=extension,
                    parent_name=parent.name,
                    relative_path=relative_path,
                    full_path=str(path),
                    size_bytes=size_bytes,
                    size_display=format_size(size_bytes),
                    modified_at=format_datetime(stat.st_mtime),
                    created_at=format_datetime(stat.st_ctime),
                )
            )
            add_ext_summary(extension, size_bytes)
            add_folder_summary(parent, "파일", size_bytes)
        except Exception as exc:
            result.errors.append({"path": str(path), "kind": "파일", "error": str(exc)})

    def add_folder(path: Path) -> None:
        try:
            name = path.name
            if should_skip_name(name):
                return

            stat = path.stat()
            parent = path.parent
            direct_file_count, direct_folder_count = count_direct_children(path)
            relative_path = os.path.relpath(path, root)

            result.rows.append(
                InventoryItem(
                    kind="폴더",
                    name=name,
                    stem=name,
                    extension="",
                    parent_name=parent.name,
                    relative_path=relative_path,
                    full_path=str(path),
                    size_bytes="",
                    size_display="",
                    modified_at=format_datetime(stat.st_mtime),
                    created_at=format_datetime(stat.st_ctime),
                    direct_file_count=direct_file_count,
                    direct_folder_count=direct_folder_count,
                )
            )
            add_folder_summary(parent, "폴더")
        except Exception as exc:
            result.errors.append({"path": str(path), "kind": "폴더", "error": str(exc)})

    log("파일/폴더 목록을 수집하는 중...")

    if include_subfolders:
        for current, dirs, files in os.walk(root):
            current_path = Path(current)
            if exclude_office_temp:
                dirs[:] = [dirname for dirname in dirs if not should_skip_name(dirname)]

            if include_folders:
                for dirname in dirs:
                    add_folder(current_path / dirname)

            if include_files:
                for filename in files:
                    add_file(current_path / filename)
    else:
        try:
            with os.scandir(root) as iterator:
                for entry in iterator:
                    try:
                        entry_path = Path(entry.path)
                        if entry.is_dir(follow_symlinks=False) and include_folders:
                            add_folder(entry_path)
                        elif entry.is_file(follow_symlinks=False) and include_files:
                            add_file(entry_path)
                    except Exception as exc:
                        result.errors.append(
                            {
                                "path": getattr(entry, "path", ""),
                                "kind": "항목",
                                "error": str(exc),
                            }
                        )
        except Exception as exc:
            result.errors.append({"path": str(root), "kind": "기준 폴더", "error": str(exc)})

    result.rows.sort(key=lambda item: (item.kind, item.relative_path.lower()))
    return result


def set_common_sheet_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def adjust_column_width(ws, max_width: int = 70) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is None:
                value_len = 0
            else:
                value_text = str(value)
                value_len = max(len(part) for part in value_text.split("\n"))
            max_len = max(max_len, value_len)

        ws.column_dimensions[column].width = min(max_len + 2, max_width)


def finalize_sheet(ws) -> None:
    set_common_sheet_style(ws)
    adjust_column_width(ws)


def save_inventory_excel(
    excel_path: str | Path,
    root_folder: str | Path,
    inventory: InventoryResult,
    options_text: str,
) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = safe_excel_sheet_title("파일_폴더_목록")
    ws.append([
        "순번",
        "구분",
        "이름",
        "이름_확장자제외",
        "확장자",
        "상위폴더명",
        "상대경로",
        "전체경로",
        "크기(Bytes)",
        "크기",
        "수정일시",
        "생성일시",
        "직접하위파일수",
        "직접하위폴더수",
    ])

    for index, item in enumerate(inventory.rows, start=1):
        ws.append([
            index,
            item.kind,
            item.name,
            item.stem,
            item.extension,
            item.parent_name,
            item.relative_path,
            item.full_path,
            item.size_bytes,
            item.size_display,
            item.modified_at,
            item.created_at,
            item.direct_file_count,
            item.direct_folder_count,
        ])
    finalize_sheet(ws)
    ws.column_dimensions["H"].width = 60

    ws_ext = wb.create_sheet(safe_excel_sheet_title("확장자_요약"))
    ws_ext.append(["확장자", "파일수", "총용량(Bytes)", "총용량"])
    for key in sorted(inventory.ext_summary):
        item = inventory.ext_summary[key]
        total_size = item["total_size_bytes"]
        ws_ext.append([item["extension"], item["file_count"], total_size, format_size(total_size)])
    finalize_sheet(ws_ext)

    ws_folder = wb.create_sheet(safe_excel_sheet_title("폴더_요약"))
    ws_folder.append(["상대경로", "전체경로", "파일수", "폴더수", "총용량(Bytes)", "총용량"])
    for key in sorted(inventory.folder_summary):
        item = inventory.folder_summary[key]
        total_size = item["total_size_bytes"]
        ws_folder.append([
            item["relative_path"],
            item["full_path"],
            item["file_count"],
            item["folder_count"],
            total_size,
            format_size(total_size),
        ])
    finalize_sheet(ws_folder)
    ws_folder.column_dimensions["B"].width = 60

    ws_errors = wb.create_sheet(safe_excel_sheet_title("오류_기록"))
    ws_errors.append(["경로", "구분", "오류내용"])
    for error in inventory.errors:
        ws_errors.append([error.get("path", ""), error.get("kind", ""), error.get("error", "")])
    finalize_sheet(ws_errors)
    ws_errors.column_dimensions["A"].width = 60

    ws_info = wb.create_sheet(safe_excel_sheet_title("실행_정보"))
    ws_info.append(["항목", "내용"])
    ws_info.append(["기준 폴더", str(Path(root_folder).resolve())])
    ws_info.append(["작성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_info.append(["옵션", options_text])
    ws_info.append(["목록 건수", len(inventory.rows)])
    ws_info.append(["오류 건수", len(inventory.errors)])
    finalize_sheet(ws_info)
    ws_info.column_dimensions["B"].width = 80

    wb.save(excel_path)
