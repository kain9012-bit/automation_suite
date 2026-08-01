from __future__ import annotations

import csv
import re
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries


MODE_WORKBOOK_SHEETS_TO_ONE = "workbook_sheets_to_one"
MODE_FILES_TO_SHEETS = "files_to_sheets"
MODE_SAME_SHEETS_TOGETHER = "same_sheets_together"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


@dataclass
class MergeOptions:
    mode: str
    source_paths: list[Path]
    output_path: Path
    add_source_file: bool = False
    add_source_sheet: bool = False
    include_hidden_sheets: bool = False


@dataclass
class MergeResult:
    output_path: Path
    sheet_count: int = 0
    data_row_count: int = 0
    skipped_count: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class SheetData:
    file_path: Path
    sheet_name: str
    rows: list[list]
    header_row: int


def is_supported_file(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def safe_file_name(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", str(name or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"[\.\s]+$", "", cleaned)
    return cleaned or "result"


def safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name or "Sheet1")).strip()
    return (cleaned or "Sheet1")[:31]


def unique_sheet_title(base: str, used: set[str]) -> str:
    safe = safe_sheet_title(base)
    candidate = safe
    index = 1
    while candidate.lower() in used:
        index += 1
        suffix = f"({index})"
        candidate = f"{safe[:31 - len(suffix)]}{suffix}"
    used.add(candidate.lower())
    return candidate


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def is_empty_row(row: list) -> bool:
    return all(normalize_value(value) == "" for value in row)


def trim_trailing_empty(row: list) -> list:
    values = list(row)
    while values and normalize_value(values[-1]) == "":
        values.pop()
    return values


def _read_csv_rows(path: Path) -> list[list[str]]:
    encodings = ("utf-8-sig", "cp949", "euc-kr")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as fp:
                return [row for row in csv.reader(fp)]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def load_source_workbook(path: Path, data_only: bool = False):
    ext = path.suffix.lower()
    if ext == ".xls":
        raise ValueError(".xls 파일은 현재 Python 도구에서 직접 지원하지 않습니다. .xlsx로 저장한 뒤 사용해 주세요.")
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("지원하는 파일 형식은 .xlsx, .xlsm, .csv 입니다.")

    if ext == ".csv":
        wb = Workbook()
        ws = wb.active
        ws.title = safe_sheet_title(path.stem)
        for row in _read_csv_rows(path):
            ws.append(row)
        return wb, True

    return load_workbook(path, data_only=data_only), False


def worksheet_to_rows(ws) -> list[list]:
    rows: list[list] = []
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, min_col=1, max_col=max_col, values_only=True):
        rows.append([normalize_value(value) for value in row])
    return rows


def detect_header_row_from_rows(rows: list[list], max_scan_rows: int = 50) -> int:
    best_index = 1
    best_score = -1
    for idx, row in enumerate(rows[:max_scan_rows], start=1):
        values = [normalize_value(value) for value in row]
        nonempty = [value for value in values if value != ""]
        if not nonempty:
            continue
        text_count = sum(1 for value in nonempty if isinstance(value, str) and not re.fullmatch(r"[\d,.%-]+", value))
        score = len(nonempty) * 3 + text_count
        if score > best_score:
            best_score = score
            best_index = idx
    return best_index


def collect_sheet_data(path: Path, include_hidden_sheets: bool = False) -> list[SheetData]:
    wb_values, _ = load_source_workbook(path, data_only=True)
    collected: list[SheetData] = []
    for ws in wb_values.worksheets:
        if not include_hidden_sheets and ws.sheet_state != "visible":
            continue
        rows = worksheet_to_rows(ws)
        if not rows:
            continue
        header_row = detect_header_row_from_rows(rows)
        collected.append(SheetData(path, ws.title, rows, header_row))
    return collected


def _copy_cell(src, dst) -> None:
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    dst.number_format = src.number_format
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def copy_worksheet_full(ws_in, wb_out: Workbook, title: str):
    ws_out = wb_out.create_sheet(title)
    for col_idx in range(1, (ws_in.max_column or 1) + 1):
        letter = get_column_letter(col_idx)
        src_dim = ws_in.column_dimensions.get(letter)
        if src_dim is not None:
            dst_dim = ws_out.column_dimensions[letter]
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
            dst_dim.outlineLevel = src_dim.outlineLevel

    for row_idx in range(1, (ws_in.max_row or 0) + 1):
        src_dim = ws_in.row_dimensions.get(row_idx)
        if src_dim is not None:
            dst_dim = ws_out.row_dimensions[row_idx]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden
            dst_dim.outlineLevel = src_dim.outlineLevel
        for col_idx in range(1, (ws_in.max_column or 1) + 1):
            _copy_cell(ws_in.cell(row=row_idx, column=col_idx), ws_out.cell(row=row_idx, column=col_idx))

    for merged in list(ws_in.merged_cells.ranges):
        try:
            ws_out.merge_cells(str(merged))
        except Exception:
            pass
    if ws_in.freeze_panes:
        ws_out.freeze_panes = ws_in.freeze_panes
    return ws_out


def apply_header_style(sample_path: Path, sample_sheet_name: str, ws_out, header_row: int, width: int) -> None:
    try:
        wb_style, _ = load_source_workbook(sample_path, data_only=False)
        ws_style = wb_style[sample_sheet_name] if sample_sheet_name in wb_style.sheetnames else wb_style.worksheets[0]
        for col_idx in range(1, min(width, ws_style.max_column or width) + 1):
            letter = get_column_letter(col_idx)
            src_dim = ws_style.column_dimensions.get(letter)
            if src_dim is not None:
                ws_out.column_dimensions[letter].width = src_dim.width

        for row_idx in range(1, min(header_row, ws_style.max_row or header_row) + 1):
            src_dim = ws_style.row_dimensions.get(row_idx)
            if src_dim is not None:
                ws_out.row_dimensions[row_idx].height = src_dim.height
            for col_idx in range(1, min(width, ws_style.max_column or width) + 1):
                _copy_cell(ws_style.cell(row=row_idx, column=col_idx), ws_out.cell(row=row_idx, column=col_idx))

        for merged in list(ws_style.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged))
            if min_row <= header_row and max_row <= header_row and max_col <= width:
                try:
                    ws_out.merge_cells(str(merged))
                except Exception:
                    pass
    except Exception:
        return


def write_rows_sheet(
    wb_out: Workbook,
    title: str,
    rows: list[list],
    sample_path: Path | None = None,
    sample_sheet_name: str = "",
    header_row: int = 1,
) -> None:
    ws_out = wb_out.create_sheet(safe_sheet_title(title))
    width = max((len(row) for row in rows), default=1)
    for row_idx, row in enumerate(rows, start=1):
        for col_idx in range(1, width + 1):
            ws_out.cell(row=row_idx, column=col_idx).value = row[col_idx - 1] if col_idx <= len(row) else None
    if sample_path:
        apply_header_style(sample_path, sample_sheet_name, ws_out, header_row, width)
        for row_idx, row in enumerate(rows, start=1):
            for col_idx in range(1, width + 1):
                cell = ws_out.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = row[col_idx - 1] if col_idx <= len(row) else None


def pad_row(row: list, width: int) -> list:
    values = list(row[:width])
    if len(values) < width:
        values.extend([""] * (width - len(values)))
    return values


def build_merged_rows(sheets: list[SheetData], add_source_file: bool, add_source_sheet: bool) -> tuple[list[list], int, int]:
    if not sheets:
        return [], 1, 0
    header_row = sheets[0].header_row
    header_rows = [trim_trailing_empty(row) for row in sheets[0].rows[:header_row]]
    data_rows_by_sheet = [(sheet, sheet.rows[sheet.header_row:]) for sheet in sheets]

    width = 0
    for row in header_rows:
        width = max(width, len(trim_trailing_empty(row)))
    for _sheet, data_rows in data_rows_by_sheet:
        for row in data_rows:
            width = max(width, len(trim_trailing_empty(row)))
    width = max(width, 1)

    extra_headers: list[str] = []
    if add_source_file:
        extra_headers.append("source_file")
    if add_source_sheet:
        extra_headers.append("source_sheet")

    final_rows: list[list] = []
    for idx, row in enumerate(header_rows):
        line = pad_row(row, width)
        if extra_headers:
            line.extend(extra_headers if idx == len(header_rows) - 1 else [""] * len(extra_headers))
        final_rows.append(line)

    data_count = 0
    for sheet, data_rows in data_rows_by_sheet:
        for row in data_rows:
            if is_empty_row(row):
                continue
            line = pad_row(row, width)
            if add_source_file:
                line.append(sheet.file_path.name)
            if add_source_sheet:
                line.append(sheet.sheet_name)
            final_rows.append(line)
            data_count += 1

    return final_rows, header_row, data_count


def merge_workbook_sheets_to_one(options: MergeOptions, log: Callable[[str], None] | None = None) -> MergeResult:
    source = options.source_paths[0]
    sheets = collect_sheet_data(source, include_hidden_sheets=options.include_hidden_sheets)
    if not sheets:
        raise ValueError("병합할 보이는 시트를 찾지 못했습니다.")

    rows, header_row, data_count = build_merged_rows(
        sheets,
        add_source_file=False,
        add_source_sheet=options.add_source_sheet,
    )
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    write_rows_sheet(wb_out, "merged", rows, sheets[0].file_path, sheets[0].sheet_name, header_row)
    wb_out.save(options.output_path)
    if log:
        log(f"시트 {len(sheets)}개를 하나의 시트로 병합했습니다.")
    return MergeResult(options.output_path, sheet_count=1, data_row_count=data_count)


def merge_files_to_sheets(options: MergeOptions, log: Callable[[str], None] | None = None) -> MergeResult:
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    used: set[str] = set()
    result = MergeResult(options.output_path)

    for path in options.source_paths:
        try:
            wb_in, _ = load_source_workbook(path, data_only=False)
            ws_in = next((ws for ws in wb_in.worksheets if options.include_hidden_sheets or ws.sheet_state == "visible"), None)
            if ws_in is None:
                result.skipped_count += 1
                result.warnings.append(f"{path.name}: 보이는 시트가 없습니다.")
                continue
            title = unique_sheet_title(path.stem, used)
            copy_worksheet_full(ws_in, wb_out, title)
            result.sheet_count += 1
            if log:
                log(f"시트 추가: {path.name} -> {title}")
        except Exception as exc:
            result.skipped_count += 1
            result.warnings.append(f"{path.name}: {exc}")

    if result.sheet_count == 0:
        raise ValueError("결과 파일에 추가할 시트가 없습니다.")
    wb_out.save(options.output_path)
    return result


def merge_same_sheets_together(options: MergeOptions, log: Callable[[str], None] | None = None) -> MergeResult:
    groups: dict[str, list[SheetData]] = {}
    result = MergeResult(options.output_path)
    for path in options.source_paths:
        try:
            for sheet in collect_sheet_data(path, include_hidden_sheets=options.include_hidden_sheets):
                groups.setdefault(sheet.sheet_name, []).append(sheet)
        except Exception as exc:
            result.skipped_count += 1
            result.warnings.append(f"{path.name}: {exc}")

    if not groups:
        raise ValueError("병합할 보이는 시트를 찾지 못했습니다.")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    used: set[str] = set()
    for sheet_name, sheets in groups.items():
        rows, header_row, data_count = build_merged_rows(
            sheets,
            add_source_file=options.add_source_file,
            add_source_sheet=options.add_source_sheet,
        )
        title = unique_sheet_title(sheet_name, used)
        write_rows_sheet(wb_out, title, rows, sheets[0].file_path, sheets[0].sheet_name, header_row)
        result.sheet_count += 1
        result.data_row_count += data_count
        if log:
            log(f"같은 시트명 병합: {sheet_name} -> {title}, 데이터 {data_count}행")

    wb_out.save(options.output_path)
    return result


def merge_excel(options: MergeOptions, log: Callable[[str], None] | None = None) -> MergeResult:
    if not options.source_paths:
        raise ValueError("병합할 파일이 없습니다.")
    for path in options.source_paths:
        if not path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
        if not is_supported_file(path):
            raise ValueError(f"지원하지 않는 파일 형식입니다: {path.name}")

    options.output_path.parent.mkdir(parents=True, exist_ok=True)

    if log:
        log(f"저장 파일: {options.output_path}")
        log(f"대상 파일: {len(options.source_paths)}개")

    if options.mode == MODE_FILES_TO_SHEETS:
        return merge_files_to_sheets(options, log)
    if options.mode == MODE_SAME_SHEETS_TOGETHER:
        return merge_same_sheets_together(options, log)
    return merge_workbook_sheets_to_one(options, log)
