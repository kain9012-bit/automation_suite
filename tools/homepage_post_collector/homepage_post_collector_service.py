"""게시판 수집 엔진.

게시판 목록 페이지를 읽어 표(게시글 목록)를 찾고, 각 행의 상세 페이지에서
본문과 첨부파일 정보를 모아 엑셀로 저장한다.

설계 원칙 — 컬럼 이름에 의존하지 않는다
    게시판마다 제목 컬럼 이름이 '제목', '도구명', '자료명', '데이터명', '학교명'
    등으로 제각각이라, 헤더 문자열로 게시판을 찾으면 새 게시판이 생길 때마다
    깨진다. 그래서 '행 안에 게시글 상세 링크가 들어 있는 표'를 게시판으로 보고,
    '상세 링크가 들어 있는 칸'을 제목 칸으로 본다.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) BoardCollector/2.0"

# 게시글 상세 페이지 링크 패턴 (컬럼명과 무관하게 게시판을 식별하는 기준)
#
# 지원 솔루션 계열:
#   A. 전북교육청 표준(eGov RFC3): www.jbe.go.kr, news.jbe.go.kr, www.jbedu.kr …
#      → /board/view.jbe, /board/view.jbedu, dataSid=
#   B. 학교홈페이지형: office.jbedu.kr/{기관}, school.jbedu.kr/{기관}
#      → /{메뉴코드}/view/6855899.do (숫자 id)
#   C. 도서관형: lib.jbe.go.kr/{기관}
#      → board/view.do?…&board_idx=…  (목록의 a는 href가 비고 keyvalue 속성만 있음)
POST_LINK_PATTERNS = (
    "/board/view.",      # A 계열 (view.jbe / view.jbedu / view.do 모두 걸림)
    "/board/read.",
    "boardid=",          # 쿼리스트링에 boardId 가 있는 경우(소문자 비교)
    "datasid=",
    "board_idx=",        # C 계열 상세 링크
)
POST_LINK_REGEXES = (
    # B 계열: /MABAGAB/view/6855899.do  (신형은 .do 없이 /MABADAJ/view/6041360)
    re.compile(r"/view/\d+(?:\.do)?(?:[?#]|$)"),
)

# 목록 페이지 번호를 담는 파라미터 후보. 주소에 이미 들어 있는 것을 우선 사용한다.
PAGE_PARAM_CANDIDATES = ("startPage", "pageIndex", "pageNo", "currPage", "cpage",
                         "viewPage", "s_idx", "page")
DEFAULT_PAGE_PARAM = "startPage"

# B 계열 메뉴코드 경로: /jbetc/MABAGAB/list.do 또는 신형 /woori/MABADAJ/ 처럼
# 대문자 메뉴코드가 들어간다.
_MENU_CODE_PATH_RE = re.compile(r"/[A-Z][A-Z0-9]{3,}/(?:(?:index|list)\.do)?$")

ATTACH_EXTS = (
    ".hwp", ".hwpx", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".pdf", ".zip", ".hml", ".jpg", ".jpeg", ".png", ".gif",
)

# 엑셀에 공통으로 덧붙이는 열
EXTRA_HEADERS = ["게시판URL", "상세URL", "본문내용", "첨부파일개수", "첨부파일목록"]


class Cancelled(Exception):
    """사용자가 '중지'를 눌렀을 때."""


def http_get(session, url, **kw):
    """GET 요청. 관공서 사이트의 인증서 문제로 실패하면 검증을 완화해 한 번 더 시도."""
    try:
        return session.get(url, **kw)
    except requests.exceptions.SSLError:
        try:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        except Exception:
            pass
        return session.get(url, verify=False, **kw)


@dataclass
class CrawlResult:
    rows: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    headers: list = field(default_factory=list)
    cancelled: bool = False

    @property
    def count(self) -> int:
        return len(self.rows)

    @property
    def attach_count(self) -> int:
        return sum(int(r.get("_첨부파일개수") or 0) for r in self.rows)


# ==========================================================
# URL / 페이지
# ==========================================================

def detect_page_param(url: str) -> str:
    """주소에 이미 들어 있는 페이지 파라미터 이름을 찾는다.

    없으면 주소 모양(솔루션 계열)으로 정한다.
      · lib.jbe.go.kr 계열(board/index.do + menu_idx)  → viewPage
      · 학교홈페이지형(/{대문자 메뉴코드}/index|list.do) → s_idx
      · 그 외(전북교육청 표준)                          → startPage
    """
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    lower = {k.lower(): k for k in qs}
    for cand in PAGE_PARAM_CANDIDATES:
        if cand.lower() in lower:
            return lower[cand.lower()]

    path = parsed.path
    if "/board/index.do" in path and "menu_idx" in lower:
        return "viewPage"
    if _MENU_CODE_PATH_RE.search(path):
        return "s_idx"
    return DEFAULT_PAGE_PARAM


def build_page_url(base_url: str, page: int) -> str:
    """게시판 목록 주소에 페이지 번호를 넣은 주소를 만든다."""
    parsed = urlparse(base_url)
    qs = parse_qs(parsed.query)
    param = detect_page_param(base_url)
    if param == DEFAULT_PAGE_PARAM:
        qs["paging"] = ["ok"]
    qs[param] = [str(page)]
    return urlunparse(parsed._replace(query=urlencode(qs, doseq=True)))


_PAGE_HREF_RE = re.compile(
    r"(?:startPage|pageIndex|pageNo|currPage|cpage|viewPage|s_idx)=(\d+)", re.I)


def detect_last_page(soup: BeautifulSoup, current: int = 1) -> int:
    """페이지 이동 영역에서 마지막 페이지 번호를 추정한다."""
    best = current
    for a in soup.find_all("a"):
        href = a.get("href", "")
        # ① href의 페이지 파라미터 값 (startPage=4, s_idx=8 …)
        for m in _PAGE_HREF_RE.finditer(href):
            try:
                best = max(best, int(m.group(1)))
            except ValueError:
                pass
        # ② 숫자 텍스트 링크 (목록/페이지 이동 링크에 한정)
        text = a.get_text(strip=True)
        if text.isdigit() and ("list." in href.lower() or "page" in href.lower()):
            try:
                best = max(best, int(text))
            except ValueError:
                pass
        # ③ lib 계열: <a class="paginate_button" keyvalue="15">맨끝</a>
        kv = a.get("keyvalue")
        classes = " ".join(a.get("class") or [])
        if kv and str(kv).isdigit() and "paginate" in classes:
            best = max(best, int(kv))
    # '총 44건 (1/4 페이지)' 형태에서도 읽어 본다.
    m = re.search(r"\(\s*\d+\s*/\s*(\d+)\s*페이지\s*\)", soup.get_text(" ", strip=True))
    if m:
        try:
            best = max(best, int(m.group(1)))
        except ValueError:
            pass
    return best


# ==========================================================
# 표(게시판) 찾기
# ==========================================================

def is_attachment_href(href: str) -> bool:
    """첨부파일 URL로 보이는 href인지 판단"""
    if not href:
        return False
    h = href.lower()
    if any(h.endswith(ext) for ext in ATTACH_EXTS):
        return True
    return "download" in h or "filedown" in h


def is_post_link(href: str) -> bool:
    """행 안의 a 태그가 '게시글 상세'로 가는 링크인지 판정"""
    if not href:
        return False
    h = href.strip().lower()
    if h.startswith("#") or h.startswith("javascript:"):
        return False
    if is_attachment_href(href):
        return False
    if any(p in h for p in POST_LINK_PATTERNS):
        return True
    return any(rx.search(h) for rx in POST_LINK_REGEXES)


def resolve_post_link(a, page_url: str) -> str:
    """a 태그에서 상세 페이지 URL을 얻는다. 실패 시 빈 문자열.

    · 보통: href가 상세 링크 패턴이면 절대주소로 변환.
    · lib.jbe.go.kr 계열: href가 비어 있고 keyvalue="글번호"만 있으므로,
      목록 주소(board/index.do?menu_idx=..&manage_idx=..)를 바탕으로
      board/view.do?…&board_idx=글번호 를 만들어 준다.
    """
    href = (a.get("href") or "").strip()
    if is_post_link(href):
        return urljoin(page_url, href)

    key = a.get("keyvalue")
    if key and str(key).isdigit():
        parsed = urlparse(page_url)
        if "/board/index.do" in parsed.path:
            qs = parse_qs(parsed.query)
            if "menu_idx" in {k.lower() for k in qs}:
                qs["board_idx"] = [str(key)]
                qs.pop("viewPage", None)
                new = parsed._replace(
                    path=parsed.path.replace("index.do", "view.do"),
                    query=urlencode(qs, doseq=True))
                return urlunparse(new)
    return ""


def _row_post_link(tr, page_url: str) -> str:
    """행(tr) 안에서 상세 링크 하나를 찾는다."""
    for a in tr.find_all("a"):
        url = resolve_post_link(a, page_url)
        if url:
            return url
    return ""


def count_post_links(table, page_url: str = "") -> int:
    """표 본문 안에 있는 게시글 상세 링크 개수(행 단위)."""
    body = table.find("tbody") or table
    count = 0
    for tr in body.find_all("tr"):
        if _row_post_link(tr, page_url):
            count += 1
    return count


def find_title_cell_index(table, page_url: str = "") -> int:
    """게시글 상세 링크가 들어있는 td의 컬럼 인덱스. 실패 시 -1.

    반응형 게시판은 번호 칸에도 (모바일용) 제목 링크가 숨어 있어 후보가 여럿일
    수 있다. 링크 텍스트가 가장 긴 칸을 고르고, 동률이면 뒤쪽 칸을 택한다.
    """
    body = table.find("tbody") or table
    for tr in body.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue
        candidates = []   # (텍스트 길이, 칸 번호)
        for i, td in enumerate(tds):
            for a in td.find_all("a"):
                if resolve_post_link(a, page_url):
                    text = (a.get("title") or "").strip() or a.get_text(" ", strip=True)
                    candidates.append((len(text), i))
                    break
        if candidates:
            return max(candidates)[1]
    return -1


def find_board_table(soup: BeautifulSoup, page_url: str = ""):
    """게시글 상세 링크가 가장 많이 들어 있는 표를 게시판으로 본다.

    컬럼명('제목', '도구명', '자료명', '데이터명', '학교명' …)에 의존하지 않는다.
    링크가 하나도 안 잡히는 예외적인 경우에만 헤더 키워드 방식으로 물러선다.
    """
    best_table, best_score = None, 0
    for table in soup.find_all("table"):
        score = count_post_links(table, page_url)
        if score > best_score:
            best_table, best_score = table, score
    if best_table is not None:
        return best_table

    for table in soup.find_all("table"):
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        if not headers:
            continue
        header_text = " ".join(headers)
        if "번호" in header_text and any(
            k in header_text
            for k in ["제목", "도구명", "자료명", "데이터명", "한 줄 후기", "한줄소감", "학교명"]
        ):
            return table
    return None


def collect_gallery_links(soup: BeautifulSoup, page_url: str) -> list:
    """표가 없는 갤러리/카드형 게시판에서 게시글 링크 목록을 뽑는다.

    같은 게시판(boardId가 URL에 있으면 그 게시판)의 상세 링크만 모으고,
    글번호(dataSid/board_idx) 기준으로 중복을 제거한다.
    반환: [{'title', 'url'}]
    """
    qs = parse_qs(urlparse(page_url).query)
    want_board = (qs.get("boardId") or qs.get("boardid") or [None])[0]

    out, seen = [], set()
    for a in soup.find_all("a"):
        full = resolve_post_link(a, page_url)
        if not full:
            continue
        fqs = parse_qs(urlparse(full).query)
        if want_board:
            link_board = (fqs.get("boardId") or fqs.get("boardid") or [None])[0]
            if link_board != want_board:
                continue
        m = re.search(r"/view/(\d+)(?:\.do)?(?:[?#]|$)", full)
        key = (fqs.get("dataSid") or fqs.get("board_idx")
               or ([m.group(1)] if m else None) or [full])[0]
        title = (a.get("title") or "").strip() or a.get_text(" ", strip=True)
        if not title or key in seen:
            # 썸네일(제목 없는 이미지 링크)은 건너뛰고 제목 있는 링크를 기다린다.
            continue
        seen.add(key)
        out.append({"title": title, "url": full})
    return out


def read_headers(table) -> list:
    """표의 헤더(th) 목록. 빈 헤더는 자리를 유지하되 이름을 붙여 준다."""
    headers = [th.get_text(strip=True) for th in table.find_all("th")]
    out, blank = [], 0
    for h in headers:
        if h:
            out.append(h)
        else:
            blank += 1
            out.append(f"열{blank}")
    return out


# ==========================================================
# 본문 / 첨부
# ==========================================================

def extract_body_from_view(html: str, title_hint: str) -> str:
    """상세 페이지에서 제목 아래에 있는 본문만 최대한 텍스트로 추출"""
    soup = BeautifulSoup(html, "html.parser")
    lines = [ln.strip() for ln in soup.get_text("\n", strip=True).splitlines() if ln.strip()]

    idx = None
    hint = (title_hint or "").strip()
    for i, line in enumerate(lines):
        if hint and hint in line:
            idx = i
            break
    if idx is None:
        return ""

    start = idx + 1
    while start < len(lines) and lines[start].startswith(("작성자", "작성일", "등록일", "조회수")):
        start += 1

    body_lines = []
    for line in lines[start:]:
        if line.startswith(("이전글", "다음글", "목록")):
            break
        if "공공저작물 자유이용 허락 표시" in line:
            break
        body_lines.append(line)
    return "\n".join(body_lines).strip()


def safe_filename(name: str) -> str:
    """윈도우 파일 이름에 쓸 수 없는 문자 제거"""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip() or "무제"


def clean_attachment_display_name(text: str) -> str:
    """'계획서.hwp (123KB)' → '계획서.hwp'"""
    name = (text or "").strip()
    m = re.match(r"(.+?)\s*\(\s*\d[\d,.]*\s*(kb|mb|byte|bytes)\s*\)?", name, flags=re.I)
    if m:
        name = m.group(1).strip()
    return name


def extract_attachments_from_detail(soup: BeautifulSoup, base_url: str) -> list:
    """상세 페이지에서 첨부파일 링크 목록을 추출(중복 제거)."""
    files, seen = [], set()
    for a in soup.find_all("a"):
        text = a.get_text(" ", strip=True)
        href = a.get("href") or ""
        if not href:
            continue
        lower_text = text.lower()
        full_url = urljoin(base_url, href)

        if "kb)" in lower_text or "(kb" in lower_text or is_attachment_href(href):
            name = clean_attachment_display_name(text) if text else os.path.basename(full_url)
            if full_url in seen:
                continue
            seen.add(full_url)
            files.append({"name": name or os.path.basename(full_url), "url": full_url})
    return files


def download_attachment(session, full_url, save_dir, log_func, error_list,
                        page, post_no, title, desired_filename=None):
    """첨부파일 1개 다운로드. 성공 시 저장된 파일명, 실패 시 None."""
    try:
        resp = http_get(session, full_url, timeout=30, stream=True)
        resp.raise_for_status()

        filename = desired_filename
        if not filename:
            cd = resp.headers.get("Content-Disposition", "")
            if "filename" in cd:
                m = re.search(r"filename\*?=([^;]+)", cd)
                if m:
                    part = m.group(1).strip().strip('"')
                    if "''" in part:
                        part = part.split("''", 1)[1]
                    filename = part
        if not filename:
            filename = os.path.basename(urlparse(full_url).path) or "downloaded_file"

        filename = safe_filename(filename)
        os.makedirs(save_dir, exist_ok=True)
        full_path = os.path.join(save_dir, filename)

        base_name, ext = os.path.splitext(full_path)
        idx = 1
        while os.path.exists(full_path):
            full_path = f"{base_name}({idx}){ext}"
            idx += 1

        with open(full_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        log_func(f"      · 첨부 저장: {os.path.basename(full_path)}")
        return os.path.basename(full_path)

    except Exception as e:
        msg = f"첨부파일 다운로드 실패 ({full_url}): {e}"
        error_list.append({"page": page, "post_no": post_no, "title": title, "error": msg})
        log_func(f"      ! {msg}")
        return None


# ==========================================================
# 수집 본체
# ==========================================================

def _check_cancel(cancel_event):
    if cancel_event is not None and cancel_event.is_set():
        raise Cancelled()


def _collect_detail(session, post_url, title_hint, author_hint, page, post_no,
                    result, log_func, collect_body, download_files, attach_dir,
                    cancel_event, note_error):
    """상세 페이지에서 본문·첨부 목록을 모으고(옵션) 첨부를 내려받는다.

    반환: (body, attach_names)
    """
    body, attach_names = "", []
    if not post_url or not (collect_body or download_files):
        return body, attach_names
    try:
        dresp = http_get(session, post_url, timeout=15)
        dresp.raise_for_status()
        dresp.encoding = dresp.apparent_encoding or "utf-8"
        detail_html = dresp.text
        detail_soup = BeautifulSoup(detail_html, "html.parser")

        if collect_body:
            try:
                body = extract_body_from_view(detail_html, title_hint)
            except Exception as e:
                note_error(page, post_no, title_hint, f"본문 수집 실패: {e}")
                log_func(f"    - [{title_hint}] 본문 수집 실패")

        try:
            attach_infos = extract_attachments_from_detail(detail_soup, post_url)
        except Exception as e:
            attach_infos = []
            note_error(page, post_no, title_hint, f"첨부파일 목록 추출 실패: {e}")
            log_func(f"    - [{title_hint}] 첨부파일 정보 추출 실패")

        attach_names = [i["name"] for i in attach_infos]

        if download_files and attach_dir and attach_infos:
            base_name = title_hint or f"게시글_{post_no or (len(result.rows) + 1)}"
            if author_hint:
                base_name = f"{base_name}_{author_hint}"
            post_folder = os.path.join(attach_dir, safe_filename(base_name))
            os.makedirs(post_folder, exist_ok=True)
            for info in attach_infos:
                _check_cancel(cancel_event)
                download_attachment(
                    session=session, full_url=info["url"], save_dir=post_folder,
                    log_func=log_func, error_list=result.errors, page=page,
                    post_no=post_no, title=title_hint, desired_filename=info["name"])

    except Cancelled:
        raise
    except Exception as e:
        note_error(page, post_no, title_hint, f"상세 페이지 요청 실패: {e}")
        log_func(f"    - [{title_hint}] 상세 페이지 불러오기 실패")
    return body, attach_names


def crawl_board(board_url, start_page, end_page, max_posts, log_func,
                collect_body=True, download_files=False, attach_dir=None,
                cancel_event=None, progress_func=None) -> CrawlResult:
    """게시판을 훑어 행 목록을 만든다.

    각 행(dict)은 화면에 보이는 헤더를 키로 갖고, 아래 특수 키가 덧붙는다.
        _상세URL, _본문내용, _첨부파일개수, _첨부파일목록
    """
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    result = CrawlResult()
    headers = None
    collected = 0

    def note_error(page, post_no, title, error):
        result.errors.append({"page": page, "post_no": post_no, "title": title, "error": error})

    try:
        for page in range(start_page, end_page + 1):
            _check_cancel(cancel_event)
            if max_posts and collected >= max_posts:
                break

            page_url = build_page_url(board_url, page)
            log_func(f"[페이지 {page}] 요청 중...")

            try:
                resp = http_get(session, page_url, timeout=15)
                resp.raise_for_status()
            except Exception as e:
                log_func(f"  → 페이지 요청 실패: {e}")
                note_error(page, "", "", f"페이지 요청 실패: {e}")
                continue

            resp.encoding = resp.apparent_encoding or "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")

            table = find_board_table(soup, page_url)
            if not table:
                # 갤러리/카드형(표 없음) 게시판 폴백
                links = collect_gallery_links(soup, page_url)
                if not links:
                    log_func("  → 게시판 표(또는 게시글 목록)를 찾지 못했습니다. (이 페이지는 건너뜀)")
                    note_error(page, "", "", "게시판 표 미발견")
                    continue
                if headers is None:
                    headers = ["제목"]
                    result.headers = headers
                    log_func(f"  → 목록형(갤러리) 게시판으로 인식 — 이 페이지 {len(links)}건")
                for item in links:
                    _check_cancel(cancel_event)
                    if max_posts and collected >= max_posts:
                        break
                    title_hint, post_url = item["title"], item["url"]
                    try:
                        body, attach_names = _collect_detail(
                            session, post_url, title_hint, "", page, "",
                            result, log_func, collect_body, download_files,
                            attach_dir, cancel_event, note_error)
                        row = {"제목": title_hint,
                               "_상세URL": post_url,
                               "_본문내용": body if collect_body else "",
                               "_첨부파일개수": len(attach_names),
                               "_첨부파일목록": "; ".join(attach_names) if attach_names else ""}
                        collected += 1
                        result.rows.append(row)
                        log_func(f"    - 수집됨: {title_hint}")
                        if progress_func:
                            progress_func(collected)
                    except Cancelled:
                        raise
                    except Exception as e:
                        note_error(page, "", title_hint, f"행 파싱 실패: {e}")
                        log_func(f"    - 행 파싱 중 오류 (건너뜀): {e}")
                continue

            if headers is None:
                headers = read_headers(table)
                result.headers = headers
                log_func(f"  → 인식된 헤더: {headers}")

            title_idx = find_title_cell_index(table, page_url)
            if page == start_page and 0 <= title_idx < len(headers):
                log_func(f"  → 제목 칸 인식: '{headers[title_idx]}' ({title_idx + 1}번째)")

            tbody = table.find("tbody") or table
            for tr in tbody.find_all("tr"):
                _check_cancel(cancel_event)
                if max_posts and collected >= max_posts:
                    break

                tds = tr.find_all("td")
                if not tds:
                    continue

                try:
                    row = {}
                    for i, h in enumerate(headers):
                        row[h] = tds[i].get_text(" ", strip=True) if i < len(tds) else ""

                    post_no = title_hint = author_hint = ""

                    # 제목: 상세 링크가 있는 칸을 최우선으로 사용
                    if 0 <= title_idx < len(tds):
                        link = next(
                            (a for a in tds[title_idx].find_all("a")
                             if resolve_post_link(a, page_url)), None)
                        if link is not None:
                            title_hint = (link.get("title") or "").strip() \
                                or link.get_text(" ", strip=True)
                        else:
                            title_hint = tds[title_idx].get_text(" ", strip=True)

                    for h in headers:
                        if not post_no and "번호" in h:
                            post_no = row.get(h, "")
                        if not title_hint and any(
                            k in h for k in ["제목", "도구명", "자료명", "데이터명",
                                             "학교명", "한 줄 후기", "한줄소감"]):
                            title_hint = row.get(h, "")
                        if not author_hint and any(
                                k in h for k in ["부서", "작성자", "학교명", "학교", "기관"]):
                            author_hint = row.get(h, "")

                    if not title_hint and len(headers) >= 2:
                        title_hint = row.get(headers[1], "")
                    if not post_no and headers:
                        post_no = row.get(headers[0], "")

                    # 상세 페이지 주소 (href형 + keyvalue형 모두 지원)
                    post_url = _row_post_link(tr, page_url)
                    if not post_url:
                        for a in tr.find_all("a"):
                            href = a.get("href", "")
                            if not href or href.strip().startswith(("#", "javascript:")):
                                continue
                            if is_attachment_href(href):
                                continue
                            post_url = urljoin(board_url, href)
                            break

                    body, attach_names = _collect_detail(
                        session, post_url, title_hint, author_hint, page, post_no,
                        result, log_func, collect_body, download_files,
                        attach_dir, cancel_event, note_error)

                    row["_상세URL"] = post_url
                    row["_본문내용"] = body if collect_body else ""
                    row["_첨부파일개수"] = len(attach_names)
                    row["_첨부파일목록"] = "; ".join(attach_names) if attach_names else ""

                    collected += 1
                    result.rows.append(row)
                    log_func(f"    - 수집됨: {title_hint or post_no or '(제목 없음)'}")
                    if progress_func:
                        progress_func(collected)

                except Cancelled:
                    raise
                except Exception as e:
                    note_error(page, "", "", f"행 파싱 실패: {e}")
                    log_func(f"    - 행 파싱 중 오류 (건너뜀): {e}")

    except Cancelled:
        result.cancelled = True
        log_func("\n※ 사용자가 중지했습니다. 그때까지 수집한 내용만 저장합니다.")

    result.headers = headers or []
    return result


# ==========================================================
# 엑셀 저장
# ==========================================================

def adjust_column_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col].width = min(max_len + 2, 60)


def save_to_excel(headers, rows, errors, excel_path, board_url) -> str:
    """게시판 헤더를 그대로 열로 쓰고, 뒤에 공통 열을 붙여 저장."""
    wb = Workbook()
    ws = wb.active
    ws.title = "게시글"
    ws.append(list(headers) + EXTRA_HEADERS)

    for row in rows:
        values = [row.get(h, "") for h in headers]
        values += [
            board_url,
            row.get("_상세URL", ""),
            row.get("_본문내용", ""),
            row.get("_첨부파일개수", ""),
            row.get("_첨부파일목록", ""),
        ]
        ws.append(values)
    adjust_column_width(ws)

    ws_err = wb.create_sheet("오류_건너뜀")
    ws_err.append(["페이지", "게시글번호", "제목", "오류내용"])
    for err in errors:
        ws_err.append([err.get("page", ""), err.get("post_no", ""),
                       err.get("title", ""), err.get("error", "")])
    adjust_column_width(ws_err)

    wb.save(excel_path)
    return excel_path


# ==========================================================
# 미리 확인(게시판 주소가 올바른지 빠르게 점검)
# ==========================================================

def probe_board(board_url: str, timeout: float = 15.0) -> dict:
    """게시판 주소를 한 번 열어 표·헤더·제목 칸·마지막 페이지를 알려 준다.

    반환: {'ok', 'message', 'headers', 'title_column', 'rows', 'last_page'}
    """
    out = {"ok": False, "message": "", "headers": [], "title_column": "",
           "rows": 0, "last_page": 1}
    try:
        session = requests.Session()
        session.headers.update({"User-Agent": USER_AGENT})
        resp = http_get(session, build_page_url(board_url, 1), timeout=timeout)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding or "utf-8"
    except Exception as e:
        out["message"] = f"주소를 열지 못했습니다: {e}"
        return out

    # 로그인이 필요한 게시판: 본문 대신 짧은 스크립트(알림+이동)만 내려온다.
    if len(resp.text) < 1000 and ("로그인" in resp.text or "login" in resp.text.lower()):
        out["message"] = ("로그인이 필요한 게시판입니다.\n"
                          "로그인 후에만 볼 수 있는 게시판은 수집할 수 없습니다.")
        return out

    page_url = build_page_url(board_url, 1)
    soup = BeautifulSoup(resp.text, "html.parser")
    table = find_board_table(soup, page_url)
    if table is None:
        # 갤러리/카드형 폴백
        links = collect_gallery_links(soup, page_url)
        if links:
            out.update({
                "ok": True,
                "headers": ["제목"],
                "title_column": "(목록형)",
                "rows": len(links),
                "last_page": detect_last_page(soup),
            })
            out["message"] = (f"목록형(갤러리) 게시판을 찾았습니다. "
                              f"이 페이지에 {out['rows']}건이 있습니다. "
                              f"마지막 페이지는 {out['last_page']}쪽으로 보입니다.")
            return out
        out["message"] = ("이 주소에서 게시판 표를 찾지 못했습니다.\n"
                          "게시판 '목록' 화면 주소인지 확인해 주세요.\n"
                          "(로그인이 필요하거나 화면을 스크립트로 그리는 게시판은 지원되지 않습니다.)")
        return out

    headers = read_headers(table)
    idx = find_title_cell_index(table, page_url)
    out.update({
        "ok": True,
        "headers": headers,
        "title_column": headers[idx] if 0 <= idx < len(headers) else "(자동)",
        "rows": count_post_links(table, page_url),
        "last_page": detect_last_page(soup),
    })
    out["message"] = (f"게시판을 찾았습니다. 제목 칸은 '{out['title_column']}'이고, "
                      f"이 페이지에 {out['rows']}건이 있습니다. "
                      f"마지막 페이지는 {out['last_page']}쪽으로 보입니다.")
    return out
