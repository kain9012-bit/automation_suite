from __future__ import annotations

import csv
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


MODE_COLUMN = "column"
MODE_SHEET = "sheet"
MODE_CHUNK = "chunk"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


@dataclass
class SheetInfo:
    name: str
    max_row: int
    max_column: int
    auto_header_row: int
    headers: list[str]


@dataclass
class WorkbookInfo:
    path: Path
    sheet_infos: list[SheetInfo]
    csv_source: bool = False


@dataclass
class SplitOptions:
    mode: str
    source_path: Path
    output_folder: Path
    sheet_name: str = ""
    header_row: int = 1
    split_column: int = 1
    rows_per_file: int = 1000
    skip_empty_key: bool = True


@dataclass
class SplitResult:
    created_count: int = 0
    output_folder: Path | None = None
    created_files: list[Path] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def is_supported_file(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def safe_file_name(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", str(name or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"[\.\s]+$", "", cleaned)
    return cleaned or "EMPTY"


def safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name or "Sheet1")).strip()
    return (cleaned or "Sheet1")[:31]


def normalize_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def unique_path(folder: Path, file_name: str) -> Path:
    target = folder / safe_file_name(file_name)
    stem = target.stem
    suffix = target.suffix
    index = 1
    while target.exists():
        index += 1
        target = folder / f"{stem}({index}){suffix}"
    return target


def _read_csv_rows(path: Path) -> list[list[str]]:
    encodings = ("utf-8-sig", "cp949", "euc-kr")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as fp:
                return [row for row in csv.reader(fp)]
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return []


def load_source_workbook(path: Path, data_only: bool = False):
    ext = path.suffix.lower()
    if ext == ".xls":
        raise ValueError(".xls 파일은 현재 Python 도구에서 직접 지원하지 않습니다. .xlsx로 저장한 뒤 사용해 주세요.")
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("지원하는 파일 형식은 .xlsx, .xlsm, .csv 입니다.")

    if ext == ".csv":
        wb = Workbook()
        ws = wb.active
        ws.title = safe_sheet_title(path.stem)
        for row in _read_csv_rows(path):
            ws.append(row)
        return wb, True

    return load_workbook(path, data_only=data_only), False


def detect_header_row(ws, max_scan_rows: int = 50) -> int:
    max_row = min(ws.max_row or 1, max_scan_rows)
    max_col = ws.max_column or 1
    best_row = 1
    best_score = -1

    for row_idx in range(1, max_row + 1):
        values = [normalize_value(ws.cell(row=row_idx, column=col).value) for col in range(1, max_col + 1)]
        nonempty = [value for value in values if value]
        if not nonempty:
            continue
        text_count = sum(1 for value in nonempty if not re.fullmatch(r"[\d,.%-]+", value))
        score = len(nonempty) * 3 + text_count
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def get_headers(ws, header_row: int) -> list[str]:
    headers: list[str] = []
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        value = normalize_value(ws.cell(row=header_row, column=col).value)
        headers.append(value or f"열 {col}")
    return headers


def analyze_workbook(path: Path) -> WorkbookInfo:
    wb, csv_source = load_source_workbook(path, data_only=True)
    sheet_infos: list[SheetInfo] = []

    for ws in wb.worksheets:
        if ws.sheet_state not in ("visible", "visible".lower()):
            continue
        header_row = detect_header_row(ws)
        sheet_infos.append(
            SheetInfo(
                name=ws.title,
                max_row=ws.max_row or 0,
                max_column=ws.max_column or 0,
                auto_header_row=header_row,
                headers=get_headers(ws, header_row),
            )
        )

    return WorkbookInfo(path=path, sheet_infos=sheet_infos, csv_source=csv_source)


def preview_rows(path: Path, sheet_name: str, max_rows: int = 80, max_cols: int = 20) -> list[list[str]]:
    wb, _ = load_source_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    rows: list[list[str]] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row or 1, max_rows),
        min_col=1,
        max_col=min(ws.max_column or 1, max_cols),
        values_only=True,
    ):
        rows.append([normalize_value(value) for value in row])
    return rows


def _copy_cell(src, dst) -> None:
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    dst.number_format = src.number_format
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def _copy_dimensions(ws_in, ws_out, row_map: dict[int, int]) -> None:
    for col_idx in range(1, ws_in.max_column + 1):
        letter = get_column_letter(col_idx)
        src_dim = ws_in.column_dimensions.get(letter)
        if src_dim is not None:
            dst_dim = ws_out.column_dimensions[letter]
            dst_dim.width = src_dim.width
            dst_dim.hidden = src_dim.hidden
            dst_dim.outlineLevel = src_dim.outlineLevel

    for old_row, new_row in row_map.items():
        src_dim = ws_in.row_dimensions.get(old_row)
        if src_dim is not None:
            dst_dim = ws_out.row_dimensions[new_row]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden
            dst_dim.outlineLevel = src_dim.outlineLevel


def _copy_merges(ws_in, ws_out, row_map: dict[int, int]) -> None:
    kept_rows = set(row_map)
    for merged in list(ws_in.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        rows = list(range(min_row, max_row + 1))
        if not all(row in kept_rows for row in rows):
            continue
        new_min_row = row_map[min_row]
        new_max_row = row_map[max_row]
        if new_max_row - new_min_row != max_row - min_row:
            continue
        start = f"{get_column_letter(min_col)}{new_min_row}"
        end = f"{get_column_letter(max_col)}{new_max_row}"
        try:
            ws_out.merge_cells(f"{start}:{end}")
        except Exception:
            pass


def copy_sheet_rows_to_file(source_path: Path, sheet_name: str, rows_to_keep: list[int], output_path: Path) -> None:
    wb_in, _ = load_source_workbook(source_path)
    ws_in = wb_in[sheet_name] if sheet_name in wb_in.sheetnames else wb_in.worksheets[0]

    wb_out = Workbook()
    default_ws = wb_out.active
    wb_out.remove(default_ws)
    ws_out = wb_out.create_sheet(safe_sheet_title(ws_in.title))

    unique_rows = sorted(set(row for row in rows_to_keep if 1 <= row <= (ws_in.max_row or 0)))
    row_map = {old_row: new_row for new_row, old_row in enumerate(unique_rows, start=1)}

    for old_row in unique_rows:
        new_row = row_map[old_row]
        for col in range(1, (ws_in.max_column or 1) + 1):
            _copy_cell(ws_in.cell(row=old_row, column=col), ws_out.cell(row=new_row, column=col))

    _copy_dimensions(ws_in, ws_out, row_map)
    _copy_merges(ws_in, ws_out, row_map)

    if ws_in.freeze_panes:
        ws_out.freeze_panes = ws_in.freeze_panes

    wb_out.save(output_path)


def split_by_sheet(options: SplitOptions, log: Callable[[str], None] | None = None) -> SplitResult:
    wb, _ = load_source_workbook(options.source_path)
    result = SplitResult(output_folder=options.output_folder)
    base = options.source_path.stem

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        try:
            target = unique_path(options.output_folder, f"{base}_{ws.title}.xlsx")
            rows = list(range(1, (ws.max_row or 0) + 1))
            copy_sheet_rows_to_file(options.source_path, ws.title, rows, target)
            result.created_files.append(target)
            result.created_count += 1
            if log:
                log(f"시트 저장: {ws.title} -> {target.name}")
        except Exception as exc:
            result.errors.append(f"{ws.title}: {exc}")

    return result


def split_by_chunk(options: SplitOptions, log: Callable[[str], None] | None = None) -> SplitResult:
    wb, _ = load_source_workbook(options.source_path)
    ws = wb[options.sheet_name] if options.sheet_name in wb.sheetnames else wb.worksheets[0]
    result = SplitResult(output_folder=options.output_folder)
    base = options.source_path.stem

    header_row = max(1, options.header_row)
    rows_per_file = max(1, options.rows_per_file)
    last_row = ws.max_row or header_row
    if last_row <= header_row:
        result.errors.append("헤더 아래에 데이터 행이 없습니다.")
        return result

    data_start = header_row + 1
    part = 1
    for start_row in range(data_start, last_row + 1, rows_per_file):
        end_row = min(start_row + rows_per_file - 1, last_row)
        rows_to_keep = [*range(1, header_row + 1), *range(start_row, end_row + 1)]
        try:
            target = unique_path(options.output_folder, f"{base}_part{part}_rows{start_row}-{end_row}.xlsx")
            copy_sheet_rows_to_file(options.source_path, ws.title, rows_to_keep, target)
            result.created_files.append(target)
            result.created_count += 1
            if log:
                log(f"행 수 기준 저장: {target.name}")
        except Exception as exc:
            result.errors.append(f"{start_row}-{end_row}: {exc}")
        part += 1

    return result


def split_by_column(options: SplitOptions, log: Callable[[str], None] | None = None) -> SplitResult:
    wb, _ = load_source_workbook(options.source_path)
    wb_values, _ = load_source_workbook(options.source_path, data_only=True)
    ws = wb[options.sheet_name] if options.sheet_name in wb.sheetnames else wb.worksheets[0]
    ws_values = wb_values[ws.title] if ws.title in wb_values.sheetnames else wb_values.worksheets[0]
    result = SplitResult(output_folder=options.output_folder)
    base = options.source_path.stem

    header_row = max(1, options.header_row)
    split_column = max(1, options.split_column)
    groups: dict[str, list[int]] = {}
    formula_cache_miss = 0

    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        formula_value = ws.cell(row=row, column=split_column).value
        key_value = ws_values.cell(row=row, column=split_column).value
        if key_value is None and isinstance(formula_value, str) and formula_value.startswith("="):
            formula_cache_miss += 1
        key = normalize_value(key_value if key_value is not None else formula_value)
        if not key and options.skip_empty_key:
            continue
        group_key = key or "EMPTY"
        groups.setdefault(group_key, []).append(row)

    if formula_cache_miss and log:
        log(
            "주의: 기준 열의 일부 수식 셀에서 계산 결과 캐시를 읽지 못했습니다. "
            "엑셀에서 파일을 열어 다시 저장한 뒤 실행하면 화면에 보이는 값 기준으로 분할됩니다."
        )

    if not groups:
        result.errors.append("분할 기준 열에 유효한 값이 없습니다.")
        return result

    for key, data_rows in groups.items():
        rows_to_keep = [*range(1, header_row + 1), *data_rows]
        try:
            target = unique_path(options.output_folder, f"{base}_{safe_file_name(key)}.xlsx")
            copy_sheet_rows_to_file(options.source_path, ws.title, rows_to_keep, target)
            result.created_files.append(target)
            result.created_count += 1
            if log:
                log(f"열 값 기준 저장: {key} -> {target.name}")
        except Exception as exc:
            result.errors.append(f"{key}: {exc}")

    return result


def split_excel(options: SplitOptions, log: Callable[[str], None] | None = None) -> SplitResult:
    if not options.source_path.exists():
        raise FileNotFoundError("원본 파일을 찾을 수 없습니다.")
    if not is_supported_file(options.source_path):
        raise ValueError("지원하는 파일 형식은 .xlsx, .xlsm, .csv 입니다.")

    options.output_folder.mkdir(parents=True, exist_ok=True)

    if log:
        log(f"원본 파일: {options.source_path}")
        log(f"저장 폴더: {options.output_folder}")

    if options.mode == MODE_SHEET:
        return split_by_sheet(options, log)
    if options.mode == MODE_CHUNK:
        return split_by_chunk(options, log)
    return split_by_column(options, log)


def copy_output_folder_path(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path
